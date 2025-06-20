import pandas as pd
import os
import unicodedata
import re
from datetime import datetime
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo

def remover_acentos(texto):
    try:
        # Normaliza o texto para decomposição NFD e remove os diacríticos
        return ''.join(c for c in unicodedata.normalize('NFD', str(texto))
                      if unicodedata.category(c) != 'Mn')
    except:
        return texto

def converter_para_inteiro(valor):
    try:
        # Primeiro tenta converter diretamente para float
        float_valor = float(valor)
        # Se não for NaN, converte para inteiro
        if not pd.isna(float_valor):
            return int(float_valor)
        return 0
    except (ValueError, TypeError):
        return 0

def criar_dataframe_novos_itens(itens_encontrados):
    # Se não houver itens, retorna None
    if not itens_encontrados:
        return None
    
    # Criar DataFrame com os novos itens
    df = pd.DataFrame(itens_encontrados)
    
    # Garantir que todas as colunas necessárias existem
    colunas_necessarias = ['DESCRIÇÃO DO ITEM', 'CÓDIGO DO ITEM', 'NÚMERO DO DOCUMENTO', 'NÚMERO PROJETO', 'ID do item']
    for coluna in colunas_necessarias:
        if coluna not in df.columns:
            df[coluna] = 0
    
    # Preencher valores NaN em todas as colunas
    for coluna in df.columns:
        if coluna in ['ID do item', 'CÓDIGO DO ITEM', 'NÚMERO DO DOCUMENTO', 'NÚMERO PROJETO']:
            # Primeiro converte para string
            df[coluna] = df[coluna].astype(str)
            # Substitui valores vazios por '0'
            df[coluna] = df[coluna].replace('', '0')
            df[coluna] = df[coluna].replace('nan', '0')
            # Converte para float, substitui NaN por 0 e converte para int
            df[coluna] = pd.to_numeric(df[coluna], errors='coerce').fillna(0).astype(int)
        else:
            df[coluna] = df[coluna].fillna('')
    
    return df

# Dicionário de categorias e palavras-chave
CATEGORIAS_PALAVRAS_CHAVE = {
    'Escalada': [
        "CINTO DE SEGURANCA", "CINTO DE SEGURANÇA", "CAPACETE ALPINISTA", "MOSQUETAO ROSCA OVAL", "MOSQUETÃO ROSCA OVAL",
        "DESCENSOR", "ASCENSOR", "TRAVA QUEDAS", "TRAVA-QUEDAS", "PROTETOR DE CORDA", "ESTRIBO",
        "CORDA S-EST", "CORDA DINAMICA", "CORDA DINÂMICA", "CORDA SEMI ESTATICA", "CORDA SEMI ESTÁTICA",
        "TALABARTE", "RIG DESCENSOR", "GANCHO PARA VARA", "BLOQUEADOR", "TRAVAQUEDA", "RETINIDA FLUTUANTE"
    ],
    'MACACÃO': ["MACACAO", "MACACOES"],
    'LUVA': ["LUVA", "LUVAS"],
    'BOTA': ["BOTA", "BOTAS"],
    'MÁSCARA': ["MASCARA"],
    'PROTETORES OURICULARES': ["PROTETOR OURICULAR", "CONCHA"]
}

def identificar_categoria(descricao):
    descricao_norm = remover_acentos(str(descricao)).upper()
    for categoria, palavras in CATEGORIAS_PALAVRAS_CHAVE.items():
        for palavra in palavras:
            if palavra in descricao_norm:
                return categoria
    return None

def verificar_item_existente(item, df_destino):
    # Usar os nomes exatos das colunas
    campos = ['DESCRIÇÃO DO ITEM', 'ITEM', 'Nº DA NF', 'Nº PROJETO']
    if df_destino.empty:
        return False
    for _, item_existente in df_destino.iterrows():
        todos_iguais = True
        for campo in campos:
            valor_novo = str(item.get(campo, '')).strip()
            valor_existente = str(item_existente.get(campo, '')).strip()
            if valor_novo != valor_existente:
                todos_iguais = False
                break
        if todos_iguais:
            return True
    return False

def processar_planilhas():
    try:
        # Arquivos
        arquivo_origem = "relatório - notas(novo).xlsx"
        arquivo_destino = "itens de escalada.xlsx"
        
        print(f"Lendo {arquivo_origem}...")
        df_origem = pd.read_excel(arquivo_origem, sheet_name="notas_com_soli")
        print(f"Total de linhas lidas: {len(df_origem)}")
        
        # Tentar ler a planilha de destino
        try:
            df_destino = pd.read_excel(arquivo_destino)
            print(f"Planilha de destino encontrada com {len(df_destino)} itens")
        except FileNotFoundError:
            print("Planilha de destino não encontrada. Criando nova planilha...")
            df_destino = pd.DataFrame()
        
        # Lista para armazenar os novos itens
        novos_itens = []
        
        # Limpar nomes de colunas para evitar problemas de espaço/acentuação
        df_origem.columns = [str(col).strip() for col in df_origem.columns]
        for idx, item in df_origem.iterrows():
            try:
                descricao = str(item.get('DESCRIÇÃO DO ITEM', '')).strip()
                codigo = item.get('ITEM', '')
                nf = item.get('Nº DA NF', '')
                projeto = item.get('Nº PROJETO', '')

                def limpar_valor(val):
                    if pd.isna(val):
                        return ''
                    return str(val).strip()
                codigo = limpar_valor(codigo)
                nf = limpar_valor(nf)
                projeto = limpar_valor(projeto)

                categoria = identificar_categoria(descricao)
                if categoria:
                    item = item.copy()
                    item['CATEGORIA'] = categoria
                    if not verificar_item_existente(item, df_destino):
                        novos_itens.append(item)
                        print(f"\n* Novo item da categoria {categoria} encontrado (linha {idx + 1}):")
                        print(f"  Descrição: {descricao}")
                        print(f"  Código: {codigo}")
                        print(f"  NF: {nf}")
                        print(f"  Projeto: {projeto}")
                    else:
                        print(f"\n* Item já existe na planilha (linha {idx + 1}):")
                        print(f"  Descrição: {descricao}")
                        print(f"  Código: {codigo}")
                        print(f"  NF: {nf}")
                        print(f"  Projeto: {projeto}")

            except KeyError as e:
                print(f"\nLinha {idx + 1} ignorada: coluna ausente: {e}")
            except Exception as e:
                print(f"\nErro ao processar linha {idx + 1}:")
                print(f"  Erro: {str(e)}")
                print("  Valores da linha:")
                for coluna in df_origem.columns:
                    print(f"    {coluna}: {item.get(coluna, '')} (tipo: {type(item.get(coluna, ''))})")
        
        # Adicionar novos itens à planilha de destino
        if novos_itens:
            df_novos = pd.DataFrame(novos_itens)
            
            # Se já existem itens, adiciona os novos
            if not df_destino.empty:
                df_final = pd.concat([df_destino, df_novos], ignore_index=True)
            else:
                df_final = df_novos
            
            # Adicionar IDs aos novos itens
            if 'ID do item' not in df_final.columns:
                df_final['ID do item'] = range(1, len(df_final) + 1)
            else:
                # Atualizar IDs apenas para os novos itens
                ultimo_id = int(df_final['ID do item'].max()) if not df_final.empty else 0
                df_final.loc[df_final['ID do item'].isna(), 'ID do item'] = range(ultimo_id + 1, ultimo_id + 1 + len(novos_itens))
            
            # Garante que a coluna CATEGORIA existe
            if 'CATEGORIA' not in df_final.columns:
                df_final['CATEGORIA'] = ''
            
            # Salvar arquivo
            print(f"\nSalvando {arquivo_destino}...")
            df_final.to_excel(arquivo_destino, index=False)
            adicionar_tabela_excel(arquivo_destino)
            print("Arquivo salvo com sucesso!")
            
            print(f"\nTotal de novos itens adicionados: {len(novos_itens)}")
            print(f"Total de itens na planilha: {len(df_final)}")
        else:
            print("\nNenhum novo item de escalada encontrado!")
        
    except Exception as e:
        print(f"Erro: {str(e)}")
        import traceback
        print("\nStack trace completo:")
        print(traceback.format_exc())

def adicionar_tabela_excel(arquivo, nome_tabela="TabelaItens"):
    wb = openpyxl.load_workbook(arquivo)
    ws = wb.worksheets[0]  # Garante que pega a primeira aba

    # Remove tabelas existentes
    if ws._tables:
        for table in ws._tables:
            ws._tables.remove(table)

    # Lê o DataFrame para obter o número exato de linhas
    df = pd.read_excel(arquivo)
    total_linhas = len(df) + 1  # +1 para incluir o cabeçalho

    # Descobre o range da tabela usando o número total de linhas
    max_col = ws.max_column
    col_letter_start = openpyxl.utils.get_column_letter(1)
    col_letter_end = openpyxl.utils.get_column_letter(max_col)
    table_range = f"{col_letter_start}1:{col_letter_end}{total_linhas}"

    # Cria a tabela
    tab = Table(displayName=nome_tabela, ref=table_range)
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)
    
    # Força o salvamento do arquivo
    wb.save(arquivo)
    wb.close()

if __name__ == "__main__":
    processar_planilhas() 
    