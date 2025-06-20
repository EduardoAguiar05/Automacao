import pandas as pd
import openpyxl

try:
    print("Método 1 - Usando pandas diretamente:")
    df1 = pd.read_excel("itens de escalada.xlsx")
    print(f"Total de linhas (método 1): {len(df1)}")
    print(f"Colunas encontradas: {df1.columns.tolist()}")
except Exception as e:
    print(f"Erro método 1: {str(e)}")

try:
    print("\nMétodo 2 - Usando openpyxl:")
    wb = openpyxl.load_workbook("itens de escalada.xlsx")
    sheet = wb.active
    print(f"Total de linhas (método 2): {sheet.max_row}")
    print(f"Total de colunas: {sheet.max_column}")
    if sheet.max_row > 0:
        print("Primeira linha:")
        for cell in sheet[1]:
            print(f"  {cell.value}")
except Exception as e:
    print(f"Erro método 2: {str(e)}")

try:
    print("\nMétodo 3 - Usando pandas com engine openpyxl:")
    df3 = pd.read_excel("itens de escalada.xlsx", engine='openpyxl')
    print(f"Total de linhas (método 3): {len(df3)}")
    print(f"Colunas encontradas: {df3.columns.tolist()}")
except Exception as e:
    print(f"Erro método 3: {str(e)}")

try:
    print("\nMétodo 4 - Verificando todas as sheets:")
    xls = pd.ExcelFile("itens de escalada.xlsx")
    print(f"Sheets encontradas: {xls.sheet_names}")
    for sheet in xls.sheet_names:
        df = pd.read_excel(xls, sheet_name=sheet)
        print(f"\nSheet '{sheet}':")
        print(f"  Total de linhas: {len(df)}")
        print(f"  Colunas: {df.columns.tolist()}")
except Exception as e:
    print(f"Erro método 4: {str(e)}") 